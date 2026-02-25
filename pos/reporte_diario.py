import tkinter as tk
import sqlite3 as sq
from datetime import datetime, timedelta
import os
import sys
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment
from tkinter import filedialog,messagebox

def resource_path(relative_path):
    
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)  
    else:
        base_path = os.path.abspath(".")  

    return os.path.join(base_path, relative_path)

class ReporteDiario():
    def __init__(self):
        
        self.ruta_bdd = resource_path('database\BDD_MINIMARKET.db')
        self.semana_actual = 1
        self.anio_actual = datetime.today().year
        self.mes_actual = datetime.today().month

        self.ventana=tk.Toplevel()
        self.ventana.title("Reporte Diario")
        self.ventana.geometry("550x600")
        self.ventana.configure(bg="#f8f8f8")
        self.ventana.resizable(False, False)
        self.ventana.focus_force()
        self.ventana.grab_set()

        for i in range(12):
            self.ventana.grid_rowconfigure(i, weight=1)
        for i in range(4):
            self.ventana.grid_columnconfigure(i, weight=1)

        self.label_titulo= tk.Label(self.ventana, text="REPORTE SEMANAL", font=('Segoe UI', 18, "bold"), bg="#f8f8f8", fg="#333")
        self.label_titulo.grid(row=0, column=0, columnspan= 4, padx= 10, pady= (20, 10), sticky="nsew")

        

        # Selector de mes
        filtro_frame = tk.Frame(self.ventana, bg="#f8f8f8")
        filtro_frame.grid(row=1, column=0, columnspan=4, pady=10)

        self.label_mes = tk.Label(filtro_frame, text="Mes:", font=('Segoe UI', 12), bg="#f8f8f8")
        self.label_mes.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="e")

        meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        self.combo_mes = tk.StringVar()
        self.combo_mes.set(meses[self.mes_actual - 1])
        self.selector_mes = tk.OptionMenu(filtro_frame, self.combo_mes, *meses, command=self.cambiar_mes)
        self.selector_mes.config(font=("Segoe UI", 11))
        self.selector_mes.grid(row=0, column=1, padx=(0, 20), pady=5, sticky="w")


        # Selector de año
        self.label_anio = tk.Label(filtro_frame, text="Año:", font=('Segoe UI', 12), bg="#f8f8f8")
        self.label_anio.grid(row=0, column=2, padx=(10, 5), pady=5, sticky="e")
        años_disponibles = [2024, 2025, 2026, 2027, 2028, 2029, 2030]
        self.combo_anio = tk.StringVar()
        self.combo_anio.set(str(self.anio_actual))
        self.selector_anio = tk.OptionMenu(filtro_frame, self.combo_anio, *años_disponibles, command=self.cambiar_anio)
        self.selector_anio.config(font=("Segoe UI", 11))
        self.selector_anio.grid(row=0, column=3, padx=(0, 10), pady=5, sticky="w")

        self.label_semana= tk.Label(self.ventana, text=f"Semana {self.semana_actual}", font=('Segoe UI', 16, "bold"), bg="#f8f8f8", fg="#555")
        self.label_semana.grid(row= 2, column= 0, columnspan= 4, padx= 10, pady= (5, 10))

        # Etiquetas de días y ventas
        self.dias_labels = {}
        dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        for i, dia in enumerate(dias):
            label_dia = tk.Label(self.ventana, text=f"{dia}:", font=('Segoe UI', 13), bg="#f8f8f8", anchor="e", width=10)
            label_dia.grid(row=i+3, column=0, padx=10, pady=5, sticky="e")

            label_venta = tk.Label(self.ventana, text="$0", font=('Segoe UI', 13, "bold"), bg="#ffffff", fg="#057A36", relief="solid", bd=1, width=12)
            label_venta.grid(row=i+3, column=1, columnspan= 2, padx=5, pady=5, sticky="w")

            self.dias_labels[dia] = label_venta

        total_semana = 0

        self.label_total_texto = tk.Label(
        self.ventana,
        text="TOTAL:",
        font=('Segoe UI', 13, "bold"),
        bg="#f8f8f8",
        anchor="e",
        width=10,
        )
        self.label_total_texto.grid(row=10, column=0, padx= 10, pady=5, sticky="e")

        self.label_total_general = tk.Label(
        self.ventana,
        text=f"${int(total_semana):,}".replace(",", "."),
        font=('Segoe UI', 13, "bold"),
        bg="#f8f8f8",
        relief="solid",
        bd=1,
        width=12,
        fg="#057A36"
        )
        self.label_total_general.grid(row=10, column=1, columnspan=2, padx=5, pady=5, sticky="w")

        #frame BOTONES
        nav_frame = tk.Frame(self.ventana, bg="#f8f8f8")
        nav_frame.grid(row=11, column=0, columnspan=4, pady=20)

        self.btn_anterior= tk.Button(nav_frame, text="<< Semana Anterior", font=('Segoe UI', 12, "bold"), bg="#e0e0e0", fg="#154360", relief="groove", command= self.semana_anterior)
        self.btn_anterior.pack(side="left", padx= 20)

        self.btn_siguiente= tk.Button(nav_frame, text="Semana Siguiente >>", font=('Segoe UI', 12, "bold"), bg="#e0e0e0", fg="#154360", relief="groove", command= self.semana_siguiente)
        self.btn_siguiente.pack(side="right", padx= 20)

        self.btn_exportar = tk.Button(self.ventana, text=" Exportar mes a Excel", font=('Segoe UI', 12, "bold"),
                                      bg="#4CAF50", fg="white", command=self.exportar_mes_excel)
        self.btn_exportar.grid(row=12, column=0, columnspan=4, padx=10, pady=10)

        self.actualizar_reporte()


    def cambiar_anio(self, selected_year):
        self.anio_actual = int(selected_year)
        self.semana_actual = 1
        self.actualizar_reporte()

    def cambiar_mes(self, selected_month):
        meses = {
            "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4, "Mayo": 5, "Junio": 6,
            "Julio": 7, "Agosto": 8, "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
        }
        self.mes_actual = meses[selected_month]
        self.semana_actual = 1
        self.actualizar_reporte()


    def semana_anterior(self):
        if self.semana_actual > 1:
            self.semana_actual -= 1
            self.actualizar_reporte()

    
    def semana_siguiente(self):
        total_semanas = len(self.obtener_semanas_5_limite(self.anio_actual, self.mes_actual))
        if self.semana_actual < total_semanas:
            self.semana_actual += 1
            self.actualizar_reporte()

        

    def obtener_semanas_5_limite(self, año, mes):
        semanas = []
        primer_dia = datetime(año, mes, 1)
        ultimo_dia = datetime(año, mes, calendar.monthrange(año, mes)[1])
        inicio = primer_dia - timedelta(days=primer_dia.weekday())
        fecha = inicio

        for _ in range(5):
            semana_inicio = fecha
            semana_fin = semana_inicio + timedelta(days=6)
            if semana_inicio > ultimo_dia:
                break
            semanas.append((semana_inicio, semana_fin))
            fecha += timedelta(weeks=1)
        return semanas

    def actualizar_reporte(self):
        semanas = self.obtener_semanas_5_limite(self.anio_actual, self.mes_actual)

        if self.semana_actual > len(semanas):
            datos_ventas = {dia: 0.0 for dia in self.dias_labels}
            self.label_semana.config(text=f"Semana {self.semana_actual} (fuera de rango)")
            return
        
        fecha_inicio, fecha_fin = semanas[self.semana_actual - 1]
        primer_dia_mes = datetime(self.anio_actual, self.mes_actual, 1)
        if fecha_inicio < primer_dia_mes:
            fecha_inicio = primer_dia_mes

        ultimo_dia_mes = datetime(self.anio_actual, self.mes_actual, calendar.monthrange(self.anio_actual, self.mes_actual)[1])
        if fecha_fin > ultimo_dia_mes:
            fecha_fin = ultimo_dia_mes


        self.label_semana.config(text=f"Semana {self.semana_actual}: {fecha_inicio.strftime('%d/%m')} - {fecha_fin.strftime('%d/%m')}")
        datos_ventas = self.obtener_ventas_por_rango_y_limitar_al_mes_actual(fecha_inicio, fecha_fin)

        total_semana = 0
        for dia, label in self.dias_labels.items():
            venta = datos_ventas.get(dia, 0.0)
            total_semana += venta
            label.config(text=f"${int(venta):,}".replace(",", "."))

        self.label_total_general.config(text=f"${int(total_semana):,}".replace(",", "."))

    def obtener_ventas_por_rango_y_limitar_al_mes_actual(self, fecha_inicio, fecha_fin):
        dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        ventas = {dia: 0.0 for dia in dias_semana}

        try:
            conn = sq.connect(self.ruta_bdd)
            cursor = conn.cursor()

            fecha_inicio_int = int(fecha_inicio.strftime('%d%m%Y'))
            fecha_fin_int = int(fecha_fin.strftime('%d%m%Y'))

            query = "SELECT Fecha_Venta, Total_Venta FROM Ventas WHERE Fecha_Venta BETWEEN ? AND ?"
            cursor.execute(query, (fecha_inicio_int, fecha_fin_int))
            filas = cursor.fetchall()

            for fecha_int, total in filas:
                dia, mes_, anio = self.descomponerFecha(fecha_int)
                fecha_obj = datetime(anio, mes_, dia)

                if fecha_obj.month == self.mes_actual and fecha_obj.year == self.anio_actual:
                    dia_nombre = self.traducir_dia(fecha_obj.strftime("%A"))
                    if dia_nombre in ventas:
                        ventas[dia_nombre] += total

            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Error al consultar la base de datos:\n{e}")
        return ventas
    
    def descomponerFecha(self, fecha_int):
        anio = fecha_int % 10000
        fecha_int //= 10000
        mes = fecha_int % 100
        fecha_int //= 100
        dia = fecha_int
        return [dia, mes, anio]

    def componerFecha(self, fecha_list):
        fecha_int = (fecha_list[0] * 1000000) + (fecha_list[1] * 10000) + (fecha_list[2])
        return fecha_int
    
    def traducir_dia(self, dia_ingles):
        traducciones = {
            "Monday": "Lunes",
            "Tuesday": "Martes",
            "Wednesday": "Miércoles",
            "Thursday": "Jueves",
            "Friday": "Viernes",
            "Saturday": "Sábado",
            "Sunday": "Domingo"
        }
        return traducciones.get(dia_ingles, dia_ingles)
    
    def exportar_mes_excel(self):
        year = self.anio_actual
        mes = self.mes_actual
        semanas = self.obtener_semanas_5_limite(year, mes)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Ventas {mes}-{year}"

        meses_es = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }

        nombre_mes = meses_es[mes]
        titulo = f"Reporte de Ventas - {nombre_mes} {year}"
        ws.merge_cells('A1:I1')
        cell_titulo = ws['A1']
        cell_titulo.value = titulo
        cell_titulo.font = Font(size=14, bold=True)
        cell_titulo.alignment = Alignment(horizontal="center")

        dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        encabezados = ["Semana"] + dias + ["Total Semana"]
        ws.append([])
        ws.append(encabezados)

        for cell in ws[3]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for i, (fecha_inicio, fecha_fin) in enumerate(semanas):
            ventas_dia = self.obtener_ventas_por_rango_y_limitar_al_mes_actual(fecha_inicio, fecha_fin)
            total_semana = sum(ventas_dia.values())
            fila = [i + 1] + [ventas_dia[d] for d in dias] + [total_semana]
            ws.append(fila)

        for col in ws.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        nombre_archivo_defecto = f"Ventas_{nombre_mes}_{year}.xlsx"
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=nombre_archivo_defecto,
                                               filetypes=[("Excel files", "*.xlsx")], title="Guardar reporte de ventas")
        if archivo:
            try:
                wb.save(archivo)
                messagebox.showinfo("Éxito", "Reporte exportado exitosamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{e}")



